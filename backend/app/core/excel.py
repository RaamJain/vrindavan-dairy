"""Utility functions for creating and managing Excel workbooks."""

from datetime import datetime
from pathlib import Path
from openpyxl import Workbook

from app.core.paths import (
    ACCOUNTS_DIR,
    MILK_RECORDS_DIR,
)

def create_accounts_workbook_if_not_exists() -> None:
    """Create the yearly accounts workbook if it does not already exist."""

    year = datetime.now().year
    filename = f"accounts_{year}.xlsx"

    workbook_path = (
        ACCOUNTS_DIR
        / filename
    )

    if workbook_path.exists():
        return

    workbook = Workbook()

    billing_sheet = workbook.active
    assert billing_sheet is not None

    billing_sheet.title = "Billing"

    payments_sheet = workbook.create_sheet(
        title="Payments",
    )

    billing_sheet.append(
        [
            "CID",
            "Customer Name",
            "Bill Month",
            "Bill Year",
            "Previous Dues",
            "Milk Amount",
            "Extras Amount",
            "Bill Amount",
            "Total Payments Received",
            "Last Payment On",
            "Current Dues",
            "Generated Date",
        ],
    )

    payments_sheet.append(
        [
            "CID",
            "Customer Name",
            "Balance Before",
            "Payment Amount",
            "Balance After",
            "Mode",
            "Date",
            "Notes",
        ],
    )

    workbook.save(
        workbook_path,
    )


def get_month_filename(
    year: int,
    month: int,
) -> str:
    """Return the workbook filename for the specified month and year."""

    return f"{year}_{month:02d}.xlsx"


def create_month_workbook_if_not_exists(
    year: int,
    month: int,
) -> None:
    """Create the monthly milk records workbook if it does not already exist."""

    filename = get_month_filename(
        year,
        month,
    )

    workbook_path = (
        MILK_RECORDS_DIR
        / filename
    )

    if workbook_path.exists():
        return

    workbook = Workbook()

    morning_sheet = workbook.active
    assert morning_sheet is not None

    morning_sheet.title = "Morning"

    evening_sheet = workbook.create_sheet(
        title="Evening",
    )

    extras_sheet = workbook.create_sheet(
        title="Extras",
    )

    headers = [
        "CID",
        "Customer Name",
        "Type",
        *(
            str(day)
            for day in range(1, 32)
        ),
        "Total",
    ]

    morning_sheet.append(
        headers,
    )

    evening_sheet.append(
        headers,
    )

    extras_sheet.append(
        [
            "CID",
            "Customer Name",
            "Date",
            "Shift",
            "Product",
            "Quantity",
        ],
    )

    workbook.save(
        workbook_path,
    )
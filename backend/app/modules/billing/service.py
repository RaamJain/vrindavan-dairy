"""Services for generating and retrieving customer billing information."""

from datetime import datetime
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import and_, or_
from sqlalchemy.orm import Session

from app.core.database import SessionLocal
from app.core.paths import ACCOUNTS_DIR
from app.modules.customers.models import Customer
from app.modules.payments.service import (
    get_customer_payment_summary,
)
from app.modules.reports.services import (
    get_customer_monthly_report,
)


def get_accounts_workbook_path(
    year: int,
):
    """Return the path to the yearly accounts workbook."""

    return (
        ACCOUNTS_DIR
        / f"accounts_{year}.xlsx"
    )


def get_bill_row(
    customer_id: int,
    month: int,
    year: int,
) -> dict[str, Any] | None:
    """Return the stored billing row for a customer.

    Args:
        customer_id: Customer identifier.
        month: Billing month.
        year: Billing year.

    Returns:
        The billing information if found, otherwise ``None``.
    """

    workbook = load_workbook(
        get_accounts_workbook_path(
            year,
        ),
    )

    billing_sheet = workbook["Billing"]

    for row in range(
        2,
        billing_sheet.max_row + 1,
    ):

        existing_customer_id = (
            billing_sheet.cell(
                row=row,
                column=1,
            ).value
        )

        existing_month = (
            billing_sheet.cell(
                row=row,
                column=3,
            ).value
        )

        existing_year = (
            billing_sheet.cell(
                row=row,
                column=4,
            ).value
        )

        if (
            existing_customer_id == customer_id
            and existing_month == month
            and existing_year == year
        ):
            return {
                "bill_amount": (
                    billing_sheet.cell(
                        row=row,
                        column=8,
                    ).value
                    or 0
                ),
                "payments_received": (
                    billing_sheet.cell(
                        row=row,
                        column=9,
                    ).value
                    or 0
                ),
                "current_dues": (
                    billing_sheet.cell(
                        row=row,
                        column=11,
                    ).value
                    or 0
                ),
                "previous_dues": (
                    billing_sheet.cell(
                        row=row,
                        column=5,
                    ).value
                    or 0
                ),
            }

    return None


def get_total_customer_payments(
    customer_id: int,
    month: int,
) -> int:
    """Return the total payments received from a customer during a month."""

    workbook = load_workbook(
        get_accounts_workbook_path(
            datetime.now().year,
        ),
    )

    payments_sheet = workbook["Payments"]

    total = 0

    for row in range(
        2,
        payments_sheet.max_row + 1,
    ):

        existing_customer_id = (
            payments_sheet.cell(
                row=row,
                column=1,
            ).value
        )

        if existing_customer_id != customer_id:
            continue

        payment_date = payments_sheet.cell(
            row=row,
            column=7,
        ).value

        if payment_date is None:
            continue

        _, payment_month, _ = map(
            int,
            payment_date.split("-"),
        )

        if payment_month == month:
            total += int(
                payments_sheet.cell(
                    row=row,
                    column=4,
                ).value
                or 0,
            )

    return total


def get_bill_preview(
    customer_id: int,
    month: int,
    year: int,
):
    """Return the complete bill preview for a customer.

    Generates or refreshes the customer's bill before assembling the
    preview information.

    Args:
        customer_id: Customer identifier.
        month: Billing month.
        year: Billing year.

    Returns:
        A dictionary containing the complete bill preview or an error
        message if the bill cannot be generated.
    """

    generate_bill(
        customer_id=customer_id,
        month=month,
        year=year,
    )

    report = get_customer_monthly_report(
        customer_id=customer_id,
        month=month,
        year=year,
    )

    if report.get("message"):
        return report

    bill_row = get_bill_row(
        customer_id,
        month,
        year,
    )

    if bill_row is None:
        return {
            "message": "Bill not generated yet",
        }

    payment_summary = get_customer_payment_summary(
        customer_id=customer_id,
    )

    latest_payment = payment_summary["latest_payment"]

    latest_payment_date = (
        latest_payment["date"]
        if latest_payment is not None
        else ""
    )

    last_payment = (
        latest_payment["payment_amount"]
        if latest_payment is not None
        else 0
    )

    customer = report["customer"]
    billing_summary = report["billing_summary"]

    return {
        "month": month,
        "year": year,
        "latest_payment": latest_payment_date,
        "last_payment": last_payment,
        "customer": customer,
        "daily_ledger": report["daily_ledger"],
        "milk_summary": report["milk_summary"],
        "extras": report["extras"],
        "monthly_total": billing_summary["monthly_total"],
        "bill_amount": bill_row["bill_amount"],
        "payments_received": get_total_customer_payments(
            customer_id,
            month,
        ),
        "previous_dues": bill_row["previous_dues"],
        "account_balance": customer["account_balance"],
    }


def generate_bill(
    customer_id: int,
    month: int,
    year: int,
) -> dict:
    """Generate or regenerate a customer's monthly bill.

    If a bill for the specified month does not exist, a new bill is created.
    Otherwise, the existing bill is updated if the billing values have changed.

    Args:
        customer_id: Unique identifier of the customer.
        month: Billing month.
        year: Billing year.

    Returns:
        A dictionary describing the generated bill or an error message.
    """

    db: Session = SessionLocal()

    try:
        customer = (
            db.query(Customer)
            .filter(
                Customer.id == customer_id,
            )
            .first()
        )

        if customer is None:
            return {
                "message": "Customer not found",
            }

        report = get_customer_monthly_report(
            customer_id=customer_id,
            month=month,
            year=year,
        )

        if report.get("message"):
            return {
                "message": "No available record found",
            }

        milk_amount = report["billing_summary"]["milk_amount"]
        extras_amount = report["billing_summary"]["extras_amount"]
        bill_amount = milk_amount + extras_amount

        workbook_path = get_accounts_workbook_path(
            year,
        )

        workbook = load_workbook(
            workbook_path,
        )

        billing_sheet = workbook["Billing"]

        existing_row = None

        for row in range(
            2,
            billing_sheet.max_row + 1,
        ):
            existing_customer_id = (
                billing_sheet.cell(
                    row=row,
                    column=1,
                ).value
            )

            existing_bill_month = (
                billing_sheet.cell(
                    row=row,
                    column=3,
                ).value
            )

            existing_bill_year = (
                billing_sheet.cell(
                    row=row,
                    column=4,
                ).value
            )

            if (
                existing_customer_id == customer_id
                and existing_bill_month == month
                and existing_bill_year == year
            ):
                existing_row = row
                break

        generated_date = datetime.now().strftime(
            "%Y-%m-%d",
        )

        # -------------------------------------------------
        # Create Bill
        # -------------------------------------------------

        if existing_row is None:

            previous_dues = float(
                customer.account_balance,
            )

            total_payments_received = 0
            last_payment_date = ""

            current_dues = (
                previous_dues
                + bill_amount
            )

            billing_sheet.append(
                [
                    customer.id,
                    customer.name,
                    month,
                    year,
                    previous_dues,
                    milk_amount,
                    extras_amount,
                    bill_amount,
                    total_payments_received,
                    last_payment_date,
                    current_dues,
                    generated_date,
                ]
            )

            customer.account_balance = current_dues

            db.commit()

            workbook.save(
                workbook_path,
            )

            return {
                "message": "Bill generated successfully",
                "customer": customer.name,
                "month": month,
                "year": year,
                "milk_amount": milk_amount,
                "extras_amount": extras_amount,
                "previous_dues": previous_dues,
                "bill_amount": bill_amount,
                "payments_received": get_total_customer_payments(
                    customer_id,
                    month,
                ),
                "current_dues": current_dues,
                "account_balance": customer.account_balance,
            }

        # -------------------------------------------------
        # Regenerate Bill
        # -------------------------------------------------

        existing_milk_amount = (
            billing_sheet.cell(
                row=existing_row,
                column=6,
            ).value
            or 0
        )

        existing_extras_amount = (
            billing_sheet.cell(
                row=existing_row,
                column=7,
            ).value
            or 0
        )

        existing_bill_amount = (
            billing_sheet.cell(
                row=existing_row,
                column=8,
            ).value
            or 0
        )

        payments_received = (
            billing_sheet.cell(
                row=existing_row,
                column=9,
            ).value
            or 0
        )

        if (
            existing_milk_amount == milk_amount
            and existing_extras_amount == extras_amount
            and existing_bill_amount == bill_amount
        ):
            current_dues = (
                billing_sheet.cell(
                    row=existing_row,
                    column=11,
                ).value
                or 0
            )

            previous_dues = (
                billing_sheet.cell(
                    row=existing_row,
                    column=5,
                ).value
                or 0
            )

            return {
                "message": "Bill already up to date",
                "customer": customer.name,
                "month": month,
                "year": year,
                "milk_amount": milk_amount,
                "extras_amount": extras_amount,
                "previous_dues": previous_dues,
                "bill_amount": bill_amount,
                "payments_received": get_total_customer_payments(
                    customer_id,
                    month,
                ),
                "current_dues": current_dues,
                "account_balance": customer.account_balance,
            }

        bill_difference = (
            float(bill_amount)
            - float(existing_bill_amount)
        )

        current_dues = (
            float(customer.account_balance)
            + bill_difference
        )

        previous_dues = (
            billing_sheet.cell(
                row=existing_row,
                column=5,
            ).value
            or 0
        )

        billing_sheet.cell(
            row=existing_row,
            column=6,
        ).value = milk_amount

        billing_sheet.cell(
            row=existing_row,
            column=7,
        ).value = extras_amount

        billing_sheet.cell(
            row=existing_row,
            column=8,
        ).value = bill_amount

        billing_sheet.cell(
            row=existing_row,
            column=11,
        ).value = current_dues

        billing_sheet.cell(
            row=existing_row,
            column=12,
        ).value = generated_date

        customer.account_balance = current_dues

        db.commit()

        workbook.save(
            workbook_path,
        )

        return {
            "message": "Bill regenerated successfully",
            "customer": customer.name,
            "month": month,
            "year": year,
            "milk_amount": milk_amount,
            "extras_amount": extras_amount,
            "previous_dues": previous_dues,
            "bill_amount": bill_amount,
            "payments_received": payments_received,
            "current_dues": current_dues,
            "account_balance": customer.account_balance,
        }

    finally:
        db.close()


def get_all_bill_previews(
    month: int,
    year: int,
) -> list[dict]:
    """Return bill previews for all customers eligible for billing.

    Active customers are always included. Inactive customers are included
    only if they have a positive outstanding account balance.

    Args:
        month: Billing month.
        year: Billing year.

    Returns:
        A list containing the bill preview for each eligible customer.
    """

    db: Session = SessionLocal()

    try:
        customers = (
            db.query(Customer)
            .filter(
                or_(
                    Customer.is_active == True,
                    and_(
                        Customer.is_active == False,
                        Customer.account_balance > 0,
                    ),
                )
            )
            .order_by(
                Customer.id,
            )
            .all()
        )

        bills: list[dict] = []

        for customer in customers:
            bill = get_bill_preview(
                customer_id=customer.id,
                month=month,
                year=year,
            )

            if bill.get("message"):
                continue

            bills.append(
                bill,
            )

        return bills

    finally:
        db.close()

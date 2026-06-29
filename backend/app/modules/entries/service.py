"""Service functions for managing daily milk entries."""

from datetime import datetime

from openpyxl import load_workbook

from app.core.paths import MILK_RECORDS_DIR
from app.core.excel import (
    create_accounts_workbook_if_not_exists,
    create_month_workbook_if_not_exists,
    get_month_filename,
)
from app.modules.extras.service import (
    save_extra_item,
)


def normalize(
    value: str | None,
) -> str:
    """Normalize a string for case-insensitive comparisons.

    Args:
        value: Value to normalize.

    Returns:
        A lowercase, whitespace-trimmed string.
    """

    return (value or "").strip().lower()


def save_milk_entry(
    customer_id: int,
    customer_name: str,
    milk_type: str,
    quantity: float,
    date: str,
    shift: str,
) -> dict:
    """Save a milk entry for a customer.

    Args:
        customer_id: Customer identifier.
        customer_name: Customer name.
        milk_type: Type of milk.
        quantity: Quantity supplied.
        date: Entry date in YYYY-MM-DD format.
        shift: Milk collection shift.

    Returns:
        A success message.
    """

    parsed_date = datetime.strptime(
        date,
        "%Y-%m-%d",
    )

    year = parsed_date.year
    month = parsed_date.month
    day = parsed_date.day

    create_accounts_workbook_if_not_exists()

    create_month_workbook_if_not_exists(
        year,
        month,
    )

    workbook_path = (
        MILK_RECORDS_DIR
        / get_month_filename(
            year,
            month,
        )
    )

    workbook = load_workbook(
        workbook_path,
    )

    worksheet = workbook[
        shift.title()
    ]

    target_row = None

    for row in range(
        2,
        worksheet.max_row + 1,
    ):
        existing_customer_id = worksheet.cell(
            row=row,
            column=1,
        ).value

        existing_milk_type = normalize(
            worksheet.cell(
                row=row,
                column=3,
            ).value,
        )

        if (
            existing_customer_id == customer_id
            and existing_milk_type == normalize(milk_type)
        ):
            target_row = row
            break

    if target_row is None:
        target_row = worksheet.max_row + 1

        worksheet.cell(
            row=target_row,
            column=1,
        ).value = customer_id

        worksheet.cell(
            row=target_row,
            column=2,
        ).value = customer_name.title()

        worksheet.cell(
            row=target_row,
            column=3,
        ).value = milk_type.title()

    day_column = day + 3

    worksheet.cell(
        row=target_row,
        column=day_column,
    ).value = quantity

    monthly_total = sum(
        worksheet.cell(
            row=target_row,
            column=column,
        ).value
        or 0
        for column in range(
            4,
            35,
        )
    )

    worksheet.cell(
        row=target_row,
        column=35,
    ).value = monthly_total

    workbook.save(
        workbook_path,
    )

    return {
        "message": "Milk entry saved successfully",
    }


def save_daily_milk_entry(
    customer_id: int,
    customer_name: str,
    cow_quantity: float,
    buffalo_quantity: float,
    date: str,
    shift: str,
) -> dict:
    """Save both cow and buffalo milk entries for a customer.

    Args:
        customer_id: Customer identifier.
        customer_name: Customer name.
        cow_quantity: Cow milk quantity.
        buffalo_quantity: Buffalo milk quantity.
        date: Entry date in YYYY-MM-DD format.
        shift: Milk collection shift.

    Returns:
        A success message.
    """

    save_milk_entry(
        customer_id=customer_id,
        customer_name=customer_name,
        milk_type="Cow",
        quantity=cow_quantity,
        date=date,
        shift=shift,
    )

    save_milk_entry(
        customer_id=customer_id,
        customer_name=customer_name,
        milk_type="Buffalo",
        quantity=buffalo_quantity,
        date=date,
        shift=shift,
    )

    return {
        "message": "Milk entry saved successfully",
    }


def get_previous_entry(
    customer_id: int,
    milk_type: str,
    date: str,
    shift: str,
) -> float:
    """Retrieve the most recent previous milk entry for a customer.

    Args:
        customer_id: Customer identifier.
        milk_type: Type of milk.
        date: Entry date in YYYY-MM-DD format.
        shift: Milk collection shift.

    Returns:
        The previous recorded milk quantity, or 0 if no previous
        entry exists.
    """

    parsed_date = datetime.strptime(
        date,
        "%Y-%m-%d",
    )

    year = parsed_date.year
    month = parsed_date.month
    day = parsed_date.day

    workbook_path = (
        MILK_RECORDS_DIR
        / get_month_filename(
            year,
            month,
        )
    )

    if not workbook_path.exists():
        return 0

    workbook = load_workbook(
        workbook_path,
    )

    worksheet = workbook[
        shift.title()
    ]

    target_row = None

    for row in range(
        2,
        worksheet.max_row + 1,
    ):
        existing_customer_id = worksheet.cell(
            row=row,
            column=1,
        ).value

        existing_milk_type = normalize(
            worksheet.cell(
                row=row,
                column=3,
            ).value,
        )

        if (
            existing_customer_id == customer_id
            and existing_milk_type == normalize(milk_type)
        ):
            target_row = row
            break

    if target_row is None:
        return 0

    for previous_day in range(
        day - 1,
        0,
        -1,
    ):
        day_column = previous_day + 3

        value = worksheet.cell(
            row=target_row,
            column=day_column,
        ).value

        if value is not None:
            return value

    return 0


def get_autofill_data(
    customer_id: int,
    date: str,
    shift: str,
) -> dict:
    """Retrieve previous milk quantities for autofill.

    Args:
        customer_id: Customer identifier.
        date: Entry date in YYYY-MM-DD format.
        shift: Milk collection shift.

    Returns:
        A dictionary containing the previous cow and buffalo milk
        quantities.
    """

    cow_quantity = get_previous_entry(
        customer_id=customer_id,
        milk_type="Cow",
        date=date,
        shift=shift,
    )

    buffalo_quantity = get_previous_entry(
        customer_id=customer_id,
        milk_type="Buffalo",
        date=date,
        shift=shift,
    )

    return {
        "cow": cow_quantity,
        "buffalo": buffalo_quantity,
    }


def save_daily_entry(
    customer_id: int,
    customer_name: str,
    date: str,
    shift: str,
    cow_quantity: float,
    buffalo_quantity: float,
    extras: list,
) -> dict:
    """Save a customer's complete daily entry.

    Args:
        customer_id: Customer identifier.
        customer_name: Customer name.
        date: Entry date in YYYY-MM-DD format.
        shift: Milk collection shift.
        cow_quantity: Cow milk quantity.
        buffalo_quantity: Buffalo milk quantity.
        extras: Additional dairy products supplied.

    Returns:
        A success message.
    """

    save_daily_milk_entry(
        customer_id=customer_id,
        customer_name=customer_name,
        cow_quantity=cow_quantity,
        buffalo_quantity=buffalo_quantity,
        date=date,
        shift=shift,
    )

    for extra in extras:
        save_extra_item(
            customer_id=customer_id,
            customer_name=customer_name,
            product_name=extra.product_name,
            quantity=extra.quantity,
            date=date,
            shift=shift,
        )

    return {
        "message": "Entry saved successfully",
    }

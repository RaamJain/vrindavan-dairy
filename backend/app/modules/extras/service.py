from datetime import datetime

from openpyxl import load_workbook

from app.core.paths import MILK_RECORDS_DIR
from app.core.excel import (
    create_month_workbook_if_not_exists,
    get_month_filename,
)


def normalize(value):
    return (value or "").strip().lower()


def save_extra_item(
    customer_id: int,
    customer_name: str,
    product_name: str,
    quantity: float,
    date: str,
    shift: str,
):
    if quantity == 0:
        return

    parsed_date = datetime.strptime(
        date,
        "%Y-%m-%d",
    )

    year = parsed_date.year
    month = parsed_date.month
    day = parsed_date.day

    create_month_workbook_if_not_exists(
        year,
        month,
    )

    workbook_path = (
        MILK_RECORDS_DIR
        / get_month_filename(
            year,
            month,
        )
    )

    workbook = load_workbook(
        workbook_path
    )

    worksheet = workbook[
        "Extras"
    ]

    target_date = (
        f"{day:02d}-{month:02d}-{year}"
    )

    target_row = None

    for row in range(
        2,
        worksheet.max_row + 1,
    ):
        existing_customer_id = worksheet.cell(
            row=row,
            column=1,
        ).value

        existing_date = worksheet.cell(
            row=row,
            column=3,
        ).value

        existing_shift = normalize(
            worksheet.cell(
                row=row,
                column=4,
            ).value
        )

        existing_product = normalize(
            worksheet.cell(
                row=row,
                column=5,
            ).value
        )

        if (
            existing_customer_id == customer_id
            and existing_date == target_date
            and existing_shift == normalize(shift)
            and existing_product
            == normalize(product_name)
        ):
            target_row = row
            break

    if target_row is None:
        target_row = (
            worksheet.max_row + 1
        )

    worksheet.cell(
        row=target_row,
        column=1,
    ).value = customer_id

    worksheet.cell(
        row=target_row,
        column=2,
    ).value = customer_name.title()

    worksheet.cell(
        row=target_row,
        column=3,
    ).value = target_date

    worksheet.cell(
        row=target_row,
        column=4,
    ).value = shift.title()

    worksheet.cell(
        row=target_row,
        column=5,
    ).value = product_name.title()

    worksheet.cell(
        row=target_row,
        column=6,
    ).value = quantity

    workbook.save(
        workbook_path
    )

    return {
        "message": "Extra item saved successfully"
    }

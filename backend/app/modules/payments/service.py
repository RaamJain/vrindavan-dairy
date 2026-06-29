from datetime import datetime

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.database import (
    SessionLocal,
)

from app.core.paths import (
    ACCOUNTS_DIR,
)

from app.modules.customers.models import (
    Customer,
)


def get_payment_history(
    customer_id: int,
):
    workbook = load_workbook(
        get_accounts_workbook_path()
    )

    payments_sheet = workbook[
        "Payments"
    ]

    history = []

    for row in range(
        2,
        payments_sheet.max_row + 1,
    ):

        existing_customer_id = (
            payments_sheet.cell(
                row=row,
                column=1,
            ).value
        )

        if (
            existing_customer_id
            != customer_id
        ):
            continue

        history.append(
            {
                "balance_before":
                    payments_sheet.cell(
                        row=row,
                        column=3,
                    ).value,

                "payment_amount":
                    payments_sheet.cell(
                        row=row,
                        column=4,
                    ).value,

                "balance_after":
                    payments_sheet.cell(
                        row=row,
                        column=5,
                    ).value,

                "mode":
                    payments_sheet.cell(
                        row=row,
                        column=6,
                    ).value,

                "date":
                    payments_sheet.cell(
                        row=row,
                        column=7,
                    ).value,

                "notes":
                    payments_sheet.cell(
                        row=row,
                        column=8,
                    ).value,
            }
        )

    history.reverse()

    return history


def get_customer_payment_summary(
    customer_id: int,
):
    db: Session = SessionLocal()

    try:

        customer = (
            db.query(Customer)
            .filter(
                Customer.id
                == customer_id
            )
            .first()
        )

        if not customer:
            return {
                "message":
                    "Customer not found"
            }

        workbook = load_workbook(
            get_accounts_workbook_path()
        )

        payments_sheet = workbook[
            "Payments"
        ]

        latest_payment = None

        for row in range(
            2,
            payments_sheet.max_row + 1,
        ):

            existing_customer_id = (
                payments_sheet.cell(
                    row=row,
                    column=1,
                ).value
            )

            if (
                existing_customer_id
                != customer_id
            ):
                continue

            latest_payment = {
                "balance_before":
                    payments_sheet.cell(
                        row=row,
                        column=3,
                    ).value,

                "payment_amount":
                    payments_sheet.cell(
                        row=row,
                        column=4,
                    ).value,

                "balance_after":
                    payments_sheet.cell(
                        row=row,
                        column=5,
                    ).value,

                "mode":
                    payments_sheet.cell(
                        row=row,
                        column=6,
                    ).value,

                "date":
                    payments_sheet.cell(
                        row=row,
                        column=7,
                    ).value,

                "notes":
                    payments_sheet.cell(
                        row=row,
                        column=8,
                    ).value,
            }

        return {
            "customer_id":
                customer.id,

            "customer_name":
                customer.name,

            "account_balance":
                float(
                    customer.account_balance
                ),

            "latest_payment":
                latest_payment,
        }

    finally:
        db.close()


def get_accounts_workbook_path():

    year = datetime.now().year

    return (
        ACCOUNTS_DIR
        / f"accounts_{year}.xlsx"
    )


def get_unpaid_bills(
    billing_sheet,
    customer_id: int,
):
    bills = []

    for row in range(
        2,
        billing_sheet.max_row + 1,
    ):

        existing_customer_id = (
            billing_sheet.cell(
                row=row,
                column=1,
            ).value
        )

        if (
            existing_customer_id
            != customer_id
        ):
            continue

        current_dues = (
            billing_sheet.cell(
                row=row,
                column=11,
            ).value
            or 0
        )

        if current_dues <= 0:
            continue

        bill_month = (
            billing_sheet.cell(
                row=row,
                column=3,
            ).value
        )

        bill_year = (
            billing_sheet.cell(
                row=row,
                column=4,
            ).value
        )

        bills.append(
            {
                "row": row,
                "month": bill_month,
                "year": bill_year,
                "current_dues": current_dues,
            }
        )

    bills.sort(
        key=lambda bill: (
            bill["year"],
            bill["month"],
        )
    )

    return bills


def record_payment(
    customer_id: int,
    payment_amount: float,
    mode: str,
    notes: str,
):
    db: Session = SessionLocal()

    try:

        customer = (
            db.query(Customer)
            .filter(
                Customer.id
                == customer_id
            )
            .first()
        )

        if not customer:
            return {
                "message":
                    "Customer not found"
            }

        if payment_amount <= 0:
            return {
                "message":
                    (
                        "Payment amount "
                        "must be greater "
                        "than zero"
                    )
            }

        workbook = load_workbook(
            get_accounts_workbook_path()
        )

        billing_sheet = workbook[
            "Billing"
        ]

        payments_sheet = workbook[
            "Payments"
        ]

        balance_before = float(
            customer.account_balance
        )

        remaining_payment = (
            payment_amount
        )

        unpaid_bills = (
            get_unpaid_bills(
                billing_sheet,
                customer_id,
            )
        )

        payment_date = (
            datetime.now().strftime(
                "%Y-%m-%d"
            )
        )

        for bill in unpaid_bills:

            if (
                remaining_payment
                <= 0
            ):
                break

            row = bill["row"]

            current_dues = (
                billing_sheet.cell(
                    row=row,
                    column=11,
                ).value
                or 0
            )

            total_payments_received = (
                billing_sheet.cell(
                    row=row,
                    column=9,
                ).value
                or 0
            )

            if (
                remaining_payment
                >= current_dues
            ):

                billing_sheet.cell(
                    row=row,
                    column=9,
                ).value = (
                    total_payments_received
                    + current_dues
                )

                billing_sheet.cell(
                    row=row,
                    column=10,
                ).value = (
                    payment_date
                )

                billing_sheet.cell(
                    row=row,
                    column=11,
                ).value = 0

                remaining_payment -= (
                    current_dues
                )

            else:

                billing_sheet.cell(
                    row=row,
                    column=9,
                ).value = (
                    total_payments_received
                    + remaining_payment
                )

                billing_sheet.cell(
                    row=row,
                    column=10,
                ).value = (
                    payment_date
                )

                billing_sheet.cell(
                    row=row,
                    column=11,
                ).value = (
                    current_dues
                    - remaining_payment
                )

                remaining_payment = 0

        balance_after = (
            balance_before
            - payment_amount
        )

        customer.account_balance = (
            balance_after
        )

        payments_sheet.append(
            [
                customer.id,

                customer.name,

                balance_before,

                payment_amount,

                balance_after,

                mode,

                payment_date,

                notes,
            ]
        )

        db.commit()

        workbook.save(
            get_accounts_workbook_path()
        )

        return {
            "message":
                (
                    "Payment recorded "
                    "successfully"
                ),

            "customer":
                customer.name,

            "balance_before":
                balance_before,

            "payment_amount":
                payment_amount,

            "balance_after":
                balance_after,
        }

    finally:
        db.close()
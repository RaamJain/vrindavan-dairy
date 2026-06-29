from email import message
from openpyxl import load_workbook
from collections import defaultdict
from decimal import Decimal
from sqlalchemy.orm import Session
from app.core.database import SessionLocal

from app.modules.customers.models import Customer
from app.modules.products.models import Product

from app.core.constants import (
    COW_MILK_RATE,
    BUFFALO_MILK_RATE,
)
from app.core.excel import (
    get_month_filename,
)
from app.core.paths import MILK_RECORDS_DIR


def normalize(value):
    return (value or "").strip().lower()


def get_customer(
    customer_id: int,
):
    db: Session = SessionLocal()

    try:
        return (
            db.query(Customer)
            .filter(
                Customer.id
                == customer_id
            )
            .first()
        )
    finally:
        db.close()
        

def find_customer_row(
    worksheet,
    customer_id: int,
    milk_type: str,
):
    for row in range(
        2,
        worksheet.max_row + 1,
    ):
        existing_customer_id = (
            worksheet.cell(
                row=row,
                column=1,
            ).value
        )

        existing_milk_type = normalize(
            worksheet.cell(
                row=row,
                column=3,
            ).value
        )

        if (
            existing_customer_id
            == customer_id
            and existing_milk_type
            == normalize(milk_type)
        ):
            return row

    return None


def get_milk_totals(
    morning_sheet,
    evening_sheet,
    customer_id: int,
):
    morning_cow_row = (
        find_customer_row(
            morning_sheet,
            customer_id,
            "Cow",
        )
    )

    morning_buffalo_row = (
        find_customer_row(
            morning_sheet,
            customer_id,
            "Buffalo",
        )
    )

    evening_cow_row = (
        find_customer_row(
            evening_sheet,
            customer_id,
            "Cow",
        )
    )

    evening_buffalo_row = (
        find_customer_row(
            evening_sheet,
            customer_id,
            "Buffalo",
        )
    )

    def total(sheet, row):
        if row is None:
            return 0

        return (
            sheet.cell(
                row=row,
                column=35,
            ).value
            or 0
        )

    return {
        "morning_cow":
            total(
                morning_sheet,
                morning_cow_row,
            ),

        "morning_buffalo":
            total(
                morning_sheet,
                morning_buffalo_row,
            ),

        "evening_cow":
            total(
                evening_sheet,
                evening_cow_row,
            ),

        "evening_buffalo":
            total(
                evening_sheet,
                evening_buffalo_row,
            ),
    }


def get_daily_milk_ledger(
    morning_sheet,
    evening_sheet,
    customer_id: int,
):
    morning_cow_row = find_customer_row(
        morning_sheet,
        customer_id,
        "Cow",
    )

    morning_buffalo_row = find_customer_row(
        morning_sheet,
        customer_id,
        "Buffalo",
    )

    evening_cow_row = find_customer_row(
        evening_sheet,
        customer_id,
        "Cow",
    )

    evening_buffalo_row = find_customer_row(
        evening_sheet,
        customer_id,
        "Buffalo",
    )

    ledger = []

    for day in range(1, 32):

        column = day + 3
        
        morning_cow = (morning_sheet.cell(
                            row=morning_cow_row,
                            column=column,
                        ).value
                        if morning_cow_row
                        else 0) or 0
        morning_buffalo = (morning_sheet.cell(
                            row=morning_buffalo_row,
                            column=column,
                        ).value
                        if morning_buffalo_row
                        else 0) or 0
        evening_cow = (evening_sheet.cell(
                            row=evening_cow_row,
                            column=column,
                        ).value
                        if evening_cow_row
                        else 0) or 0
        evening_buffalo = (evening_sheet.cell(
                            row=evening_buffalo_row,
                            column=column,
                        ).value
                        if evening_buffalo_row
                        else 0) or 0
        
        milk_total = morning_buffalo + morning_cow + evening_buffalo + evening_cow
        if milk_total == 0:
            continue    
        ledger.append(
            {
                "day": day,
                "morning_cow": morning_cow,
                "morning_buffalo": morning_buffalo,
                "evening_cow": evening_cow,
                "evening_buffalo": evening_buffalo,     
            }
        )

    return ledger


def build_milk_summary(
    milk_totals: dict,
):
    cow_total = (
        milk_totals["morning_cow"]
        + milk_totals["evening_cow"]
    )

    buffalo_total = (
        milk_totals["morning_buffalo"]
        + milk_totals["evening_buffalo"]
    )

    cow_amount = (
        cow_total
        * COW_MILK_RATE
    )

    buffalo_amount = (
        buffalo_total
        * BUFFALO_MILK_RATE
    )

    milk_amount = (
        cow_amount
        + buffalo_amount
    )

    return {
        "cow_total":
            cow_total,

        "buffalo_total":
            buffalo_total,

        "cow_rate":
            COW_MILK_RATE,

        "buffalo_rate":
            BUFFALO_MILK_RATE,

        "cow_amount":
            cow_amount,

        "buffalo_amount":
            buffalo_amount,

        "milk_amount":
            milk_amount,
    }


def get_customer_extras(
    extras_sheet,
    customer_id: int,
):
    extras = []

    for row in range(
        2,
        extras_sheet.max_row + 1,
    ):
        existing_customer_id = (
            extras_sheet.cell(
                row=row,
                column=1,
            ).value
        )

        if (
            existing_customer_id
            != customer_id
        ):
            continue

        extras.append(
            {
                "date":
                    extras_sheet.cell(
                        row=row,
                        column=3,
                    ).value,

                "shift":
                    extras_sheet.cell(
                        row=row,
                        column=4,
                    ).value,

                "product":
                    extras_sheet.cell(
                        row=row,
                        column=5,
                    ).value,

                "quantity":
                    extras_sheet.cell(
                        row=row,
                        column=6,
                    ).value,
            }
        )

    return extras


def group_extras_by_product(
    extras: list,
):
    grouped = defaultdict(
        list,
    )

    for extra in extras:
        grouped[
            extra["product"]
        ].append(
            {
                "date":
                    extra["date"],
                "shift":
                    extra["shift"],
                "quantity":
                    extra["quantity"],
            }
        )

    return grouped


def build_extras_report(
    grouped_extras,
):

    db: Session = SessionLocal()

    try:

        report = []

        for (

            product_name,

            entries,

        ) in grouped_extras.items():

            product = (

                db.query(Product)

                .filter(

                    Product.name

                    == product_name

                )

                .first()

            )

            if not product:

                continue

            total_quantity = sum(

                entry["quantity"]

                for entry in entries

            )

            amount = (

                Decimal(

                    str(total_quantity)

                )

                / Decimal("1000")

            ) * product.rate
            
            report.append(

                {

                    "product_name":

                        product.name,

                    "entries":

                        entries,
                        
                    "unit": product.unit.lower() if product.unit == "Grams" else "grams",

                    "rate":

                        float(

                            product.rate

                        ),

                    "total_quantity":

                        total_quantity,

                    "amount":

                        float(

                            amount

                        ),

                }

            )

        return report

    finally:

        db.close()


def get_extras_amount(
    extras_report: list,
):

    return sum(
        product["amount"]
        for product in extras_report
    )


def build_billing_summary(

    milk_amount: float,

    extras_amount: float,

):

    final_amount = (

        Decimal(str(milk_amount))
        + Decimal(str(extras_amount))
    )

    return {
        "milk_amount":
            float(milk_amount),
        "extras_amount":
            float(extras_amount),
        "monthly_total":
            float(final_amount),
    }
    

def get_customer_monthly_report(
    customer_id: int,
    month: int,
    year: int,
):
    workbook_path = (
        MILK_RECORDS_DIR
        / get_month_filename(
            year,
            month,
        )
    )

    if not workbook_path.exists():
        return {
            "message":
                "No records found"
        }

    workbook = load_workbook(
        workbook_path
    )

    morning_sheet = workbook[
        "Morning"
    ]

    evening_sheet = workbook[
        "Evening"
    ]

    extras_sheet = workbook[
        "Extras"
    ]
    
    milk_summary = (
        build_milk_summary(
            get_milk_totals(
            morning_sheet,
            evening_sheet,
            customer_id,
            )
        )
    )

    daily_ledger = get_daily_milk_ledger(
        morning_sheet,
        evening_sheet,
        customer_id,
    )
    
    extras = build_extras_report(group_extras_by_product(get_customer_extras(
        extras_sheet,
        customer_id,
    )))
    
    extras_amount = (
        get_extras_amount(
            extras,
        )
    )

    customer = get_customer(
        customer_id,
    )
    
    billing_summary = (

        build_billing_summary(

            milk_amount=

                milk_summary[

                    "milk_amount"

                ],

            extras_amount=

                extras_amount,

        )
    )

    return {
        "customer": {
            "id":customer.id,
            "name":customer.name,
            "contact_number": customer.contact_number,
            "address": customer.address,
            "account_balance":
                customer.account_balance,
        },
        "milk_summary": milk_summary,
        "daily_ledger": daily_ledger,
        "extras": extras,
        "billing_summary": billing_summary
    }